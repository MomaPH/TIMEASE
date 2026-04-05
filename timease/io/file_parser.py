"""Extract text content from uploaded files for use with the AI setup assistant.

Supported formats: .xlsx, .csv, .docx, .txt
Unknown formats return ("Format non supporté", "unknown").
"""
from __future__ import annotations

import csv
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def extract_content(file_path: str) -> tuple[str, str]:
    """Extract readable text from a file.

    Returns a (content, file_type) tuple where file_type is one of:
    "excel", "csv", "docx", "txt", "unknown".

    The content string uses the format::

        Feuille 'NomFeuille':
        Ligne 1: ColA=valA, ColB=valB, ...
        Ligne 2: ...

    for spreadsheet files, plain paragraph text for .docx, and raw text
    for .txt files.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        return _extract_excel_openpyxl(path)
    if suffix == ".xls":
        return _extract_excel_xlrd(path)
    if suffix == ".csv":
        return _extract_csv(path)
    if suffix in (".docx", ".doc"):
        return _extract_docx(path)
    if suffix == ".txt":
        return _extract_txt(path)
    if suffix == ".pdf":
        return _extract_pdf(path)

    return "Format non supporté", "unknown"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _extract_excel_openpyxl(path: Path) -> tuple[str, str]:
    """Read .xlsx / .xlsm workbooks via openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "openpyxl requis pour lire les fichiers Excel.", "excel"

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        logger.error("Cannot open Excel file %s: %s", path, exc)
        return f"Erreur à l'ouverture du fichier : {exc}", "excel"

    sections: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines: list[str] = [f"Feuille '{sheet_name}':"]

        headers: list[str] = []
        for cell in ws[1]:
            val = cell.value
            headers.append(str(val).strip() if val is not None else "")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue
            pairs: list[str] = []
            for col_idx, val in enumerate(row):
                if val is None:
                    continue
                header = (
                    headers[col_idx]
                    if col_idx < len(headers) and headers[col_idx]
                    else f"Col{col_idx + 1}"
                )
                pairs.append(f"{header}={val}")
            if pairs:
                lines.append(f"Ligne {row_idx}: " + ", ".join(pairs))

        sections.append("\n".join(lines))

    return "\n\n".join(sections), "excel"


def _extract_excel_xlrd(path: Path) -> tuple[str, str]:
    """Read legacy .xls workbooks via xlrd."""
    try:
        import xlrd
    except ImportError:
        return "xlrd requis pour lire les fichiers .xls. Convertissez en .xlsx si possible.", "excel"

    try:
        wb = xlrd.open_workbook(str(path))
    except Exception as exc:
        logger.error("Cannot open .xls file %s: %s", path, exc)
        return f"Erreur à l'ouverture du fichier .xls : {exc}", "excel"

    sections: list[str] = []
    for sheet in wb.sheets():
        if sheet.nrows == 0:
            continue
        lines: list[str] = [f"Feuille '{sheet.name}':"]
        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]

        for row_idx in range(1, sheet.nrows):
            pairs: list[str] = []
            for col_idx in range(sheet.ncols):
                val = sheet.cell_value(row_idx, col_idx)
                if val == "" or val is None:
                    continue
                header = headers[col_idx] if col_idx < len(headers) and headers[col_idx] else f"Col{col_idx + 1}"
                pairs.append(f"{header}={val}")
            if pairs:
                lines.append(f"Ligne {row_idx + 1}: " + ", ".join(pairs))

        sections.append("\n".join(lines))

    return "\n\n".join(sections), "excel"


def _extract_pdf(path: Path) -> tuple[str, str]:
    """Extract text from a PDF using pypdf."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return "pypdf requis pour lire les fichiers PDF.", "pdf"

    try:
        reader = PdfReader(str(path))
    except Exception as exc:
        logger.error("Cannot open PDF %s: %s", path, exc)
        return f"Erreur à l'ouverture du PDF : {exc}", "pdf"

    pages: list[str] = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        text = text.strip()
        if text:
            pages.append(f"Page {i}:\n{text}")

    if not pages:
        return "Le PDF ne contient pas de texte extractible (document scanné ?).", "pdf"

    return "\n\n".join(pages), "pdf"


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def _extract_csv(path: Path) -> tuple[str, str]:
    """Read a CSV file, auto-detecting the delimiter."""
    raw = _read_text(path)
    if raw is None:
        return f"Impossible de lire le fichier : {path.name}", "csv"

    # Auto-detect delimiter
    try:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
    rows = list(reader)

    if not rows:
        return "Fichier CSV vide.", "csv"

    headers = [h.strip() for h in rows[0]]
    lines: list[str] = [f"Feuille 'CSV':"]

    for row_idx, row in enumerate(rows[1:], start=2):
        if all(v.strip() == "" for v in row):
            continue
        pairs: list[str] = []
        for col_idx, val in enumerate(row):
            val = val.strip()
            if not val:
                continue
            header = (
                headers[col_idx]
                if col_idx < len(headers) and headers[col_idx]
                else f"Col{col_idx + 1}"
            )
            pairs.append(f"{header}={val}")
        if pairs:
            lines.append(f"Ligne {row_idx}: " + ", ".join(pairs))

    return "\n".join(lines), "csv"


# ---------------------------------------------------------------------------
# Word
# ---------------------------------------------------------------------------

def _extract_docx(path: Path) -> tuple[str, str]:
    """Extract paragraph text from a Word document."""
    try:
        from docx import Document
    except ImportError:
        return "python-docx requis pour lire les fichiers Word.", "docx"

    try:
        doc = Document(str(path))
    except Exception as exc:
        logger.error("Cannot open Word file %s: %s", path, exc)
        return f"Erreur à l'ouverture du fichier : {exc}", "docx"

    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs), "docx"


# ---------------------------------------------------------------------------
# Plain text
# ---------------------------------------------------------------------------

def _extract_txt(path: Path) -> tuple[str, str]:
    """Read a plain-text file (UTF-8, fallback latin-1)."""
    content = _read_text(path)
    if content is None:
        return f"Impossible de lire le fichier : {path.name}", "txt"
    return content, "txt"


# ---------------------------------------------------------------------------
# Shared helper
# ---------------------------------------------------------------------------

def _read_text(path: Path) -> str | None:
    """Read a file as text, trying UTF-8 first then latin-1."""
    for encoding in ("utf-8", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
        except OSError as exc:
            logger.error("Cannot read %s: %s", path, exc)
            return None
    return None
