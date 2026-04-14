"""Excel template creation and import for TIMEASE school data."""
from __future__ import annotations

import json
import logging
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from timease.engine.models import (
    Constraint,
    CurriculumEntry,
    DayConfig,
    Room,
    School,
    SchoolClass,
    SchoolData,
    SessionConfig,
    Subject,
    Teacher,
    TeacherAssignment,
    TimeslotConfig,
)

logger = logging.getLogger(__name__)

# ─── Styling helpers ───────────────────────────────────────────────────────────

_GRAY = PatternFill("solid", fgColor="CCCCCC")
_LIGHT_GRAY = PatternFill("solid", fgColor="F5F5F5")
_SECTION_GRAY = PatternFill("solid", fgColor="E0E0E0")


def _style_header(cell: Any) -> None:
    cell.fill = _GRAY
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_example(cell: Any) -> None:
    cell.fill = _LIGHT_GRAY
    cell.font = Font(italic=True, color="888888")


def _header_row(ws: Any, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        _style_header(ws.cell(row=row, column=col, value=h))
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _set_widths(ws: Any, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── Template creation ─────────────────────────────────────────────────────────

def create_template(output_path: str) -> None:
    """Create a blank Excel input template for school data.

    Args:
        output_path: Destination .xlsx path.
    """
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_mode_emploi(wb)
    _sheet_ecole(wb)
    _sheet_horaires(wb)
    _sheet_enseignants(wb)
    _sheet_classes(wb)
    _sheet_salles(wb)
    _sheet_matieres(wb)
    _sheet_affectations(wb)
    _sheet_programme(wb)
    _sheet_contraintes(wb)
    wb.save(output_path)
    logger.info("Template créé : %s", output_path)


def _sheet_mode_emploi(wb: Workbook) -> None:
    ws = wb.create_sheet("Mode d'emploi")
    ws.sheet_view.showGridLines = False

    t = ws.cell(row=1, column=1, value="TIMEASE — Guide d'utilisation du modèle Excel")
    t.font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    ws.cell(
        row=2, column=1,
        value="Remplissez chaque onglet dans l'ordre. "
              "Supprimez les lignes d'exemple (grisées) avant d'importer.",
    )
    ws.merge_cells("A2:B2")

    _style_header(ws.cell(row=4, column=1, value="Onglet"))
    _style_header(ws.cell(row=4, column=2, value="Description"))

    sheet_descriptions = [
        ("École",        "Nom de l'établissement, année académique, ville."),
        ("Horaires",     "Jours scolaires actifs, créneaux des sessions, unité de base (15/30/60 min)."),
        ("Enseignants",  "Nom, matières enseignées (virgule), heures max/semaine, indisponibilités, poids."),
        ("Classes",      "Nom de la classe, niveau, nombre d'élèves."),
        ("Salles",       "Nom, capacité, types séparés par virgule (ex : Laboratoire, Salle standard)."),
        ("Matières",     "Nom, abréviation, couleur (#RRGGBB), type de salle requis, salle nécessaire."),
        ("Affectations", "Qui enseigne quoi à quelle classe — un triplet Enseignant|Matière|Classe par ligne."),
        ("Programme",    "Volume horaire par niveau et matière (minutes/semaine + paramètres de découpage)."),
        ("Contraintes",  "Règles dur/souple. Paramètres optionnels en JSON. Tout l'onglet est optionnel."),
    ]
    for i, (name, desc) in enumerate(sheet_descriptions, 5):
        ws.cell(row=i, column=1, value=name).font = Font(bold=True)
        ws.cell(row=i, column=2, value=desc)

    ws.cell(row=16, column=1, value="Formats importants :").font = Font(bold=True)
    ws.cell(row=17, column=1, value="• Durée : 3H30 = 210 min | 2H = 120 min | 90 = 90 min")
    ws.cell(
        row=18, column=1,
        value="• Indisponibilités : 'mercredi' (jour entier) | 'lundi:Matin' (session) | "
              "'vendredi:08:00-10:00' (créneau exact). Séparateur entre entrées : point-virgule.",
    )
    ws.cell(row=19, column=1, value="• Listes de valeurs : séparées par virgule (matières, types de salles).")
    ws.cell(row=20, column=1, value="• Couleur hex : #E6F1FB — le # est requis.")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 85


def _sheet_ecole(wb: Workbook) -> None:
    ws = wb.create_sheet("École")
    _header_row(ws, ["Champ", "Valeur"])
    _set_widths(ws, [30, 40])
    fields = [
        ("Nom de l'école *", "Lycée Excellence de Dakar"),
        ("Année académique *", "2026-2027"),
        ("Ville *", "Dakar"),
    ]
    for r, (champ, ex) in enumerate(fields, 2):
        ws.cell(row=r, column=1, value=champ).font = Font(bold=True)
        _style_example(ws.cell(row=r, column=2, value=ex))


def _sheet_horaires(wb: Workbook) -> None:
    ws = wb.create_sheet("Horaires")
    _set_widths(ws, [26, 12, 15, 15])

    # ── Days section (rows 1-10) ───────────────────────────────────────────────
    sec1 = ws.cell(row=1, column=1, value="JOURS SCOLAIRES")
    sec1.font = Font(bold=True, size=11)
    sec1.fill = _SECTION_GRAY
    ws.merge_cells("A1:B1")

    for col, h in enumerate(["Jour *", "Actif *"], 1):
        _style_header(ws.cell(row=2, column=col, value=h))

    dv_yn = DataValidation(type="list", formula1='"Oui,Non"', showDropDown=False)
    ws.add_data_validation(dv_yn)

    days_defaults = [
        ("lundi", "Oui"), ("mardi", "Oui"), ("mercredi", "Oui"),
        ("jeudi", "Oui"), ("vendredi", "Oui"), ("samedi", "Oui"), ("dimanche", "Non"),
    ]
    for r, (jour, actif) in enumerate(days_defaults, 3):
        ws.cell(row=r, column=1, value=jour)
        c = ws.cell(row=r, column=2, value=actif)
        dv_yn.add(c)

    # ── Sessions section (rows 12-18) ─────────────────────────────────────────
    sec2 = ws.cell(row=12, column=1, value="SESSIONS")
    sec2.font = Font(bold=True, size=11)
    sec2.fill = _SECTION_GRAY
    ws.merge_cells("A12:C12")

    for col, h in enumerate(["Nom *", "Heure début *", "Heure fin *"], 1):
        _style_header(ws.cell(row=13, column=col, value=h))

    default_sessions = [("Matin", "08:00", "12:00"), ("Après-midi", "15:00", "17:00")]
    for r, (nom, deb, fin) in enumerate(default_sessions, 14):
        ws.cell(row=r, column=1, value=nom)
        ws.cell(row=r, column=2, value=deb)
        ws.cell(row=r, column=3, value=fin)

    # ── Base unit (row 20) ────────────────────────────────────────────────────
    b = ws.cell(row=20, column=1, value="Unité de base (minutes) *")
    b.font = Font(bold=True)
    b.fill = _SECTION_GRAY
    ws.cell(row=20, column=2, value=30)

    dv_unit = DataValidation(type="list", formula1='"15,30,60"', showDropDown=False)
    ws.add_data_validation(dv_unit)
    dv_unit.add(ws.cell(row=20, column=2))

    ws.freeze_panes = "A3"


def _sheet_enseignants(wb: Workbook) -> None:
    ws = wb.create_sheet("Enseignants")
    hdrs = [
        "Nom *", "Matières * (virgule)", "Heures max/sem. *",
        "Indisponibilités", "Poids préférence",
    ]
    _header_row(ws, hdrs)
    _set_widths(ws, [25, 45, 20, 55, 18])
    ex = ["Mme Diallo", "Mathématiques, Physique-Chimie", 18, "mercredi", 1.0]
    for col, v in enumerate(ex, 1):
        _style_example(ws.cell(row=2, column=col, value=v))


def _sheet_classes(wb: Workbook) -> None:
    ws = wb.create_sheet("Classes")
    _header_row(ws, ["Nom *", "Niveau *", "Nombre d'élèves *"])
    _set_widths(ws, [20, 20, 18])
    for col, v in enumerate(["6ème A", "6ème", 38], 1):
        _style_example(ws.cell(row=2, column=col, value=v))


def _sheet_salles(wb: Workbook) -> None:
    ws = wb.create_sheet("Salles")
    _header_row(ws, ["Nom *", "Capacité *", "Types (virgule)"])
    _set_widths(ws, [20, 12, 45])
    for col, v in enumerate(["Salle 101", 45, "Salle standard"], 1):
        _style_example(ws.cell(row=2, column=col, value=v))


def _sheet_matieres(wb: Workbook) -> None:
    ws = wb.create_sheet("Matières")
    hdrs = ["Nom *", "Abréviation *", "Couleur (hex)", "Type de salle requis", "Salle nécessaire"]
    _header_row(ws, hdrs)
    _set_widths(ws, [30, 15, 16, 25, 16])

    dv = DataValidation(type="list", formula1='"Oui,Non"', showDropDown=False)
    ws.add_data_validation(dv)
    ex = ["Mathématiques", "Maths", "#E6F1FB", "", "Oui"]
    for col, v in enumerate(ex, 1):
        _style_example(ws.cell(row=2, column=col, value=v))
    dv.add(ws.cell(row=2, column=5))


def _sheet_affectations(wb: Workbook) -> None:
    ws = wb.create_sheet("Affectations")
    _header_row(ws, ["Enseignant *", "Matière *", "Classe *"])
    _set_widths(ws, [25, 30, 20])
    for col, v in enumerate(["Mme Diallo", "Mathématiques", "6ème A"], 1):
        _style_example(ws.cell(row=2, column=col, value=v))


def _sheet_programme(wb: Workbook) -> None:
    ws = wb.create_sheet("Programme")
    hdrs = [
        "Niveau *", "Matière *", "Min/semaine *", "Mode",
        "Séances/sem.", "Min/séance", "Min min/séance", "Max min/séance",
    ]
    _header_row(ws, hdrs)
    _set_widths(ws, [15, 30, 14, 10, 14, 12, 14, 14])

    dv_mode = DataValidation(type="list", formula1='"auto,Manuel"', showDropDown=False)
    ws.add_data_validation(dv_mode)
    ex = ["6ème", "Mathématiques", 300, "auto", "", "", 50, 100]
    for col, v in enumerate(ex, 1):
        _style_example(ws.cell(row=2, column=col, value=v))
    dv_mode.add(ws.cell(row=2, column=4))


def _sheet_contraintes(wb: Workbook) -> None:
    ws = wb.create_sheet("Contraintes")
    hdrs = ["ID *", "Type *", "Catégorie *", "Description *", "Priorité", "Paramètres (JSON)"]
    _header_row(ws, hdrs)
    _set_widths(ws, [8, 10, 22, 55, 10, 45])

    dv_type = DataValidation(type="list", formula1='"dur,souple"', showDropDown=False)
    ws.add_data_validation(dv_type)
    ex = ["H1", "dur", "start_time", "Toutes les classes commencent à 08:00", 5, '{"time": "08:00"}']
    for col, v in enumerate(ex, 1):
        _style_example(ws.cell(row=2, column=col, value=v))
    dv_type.add(ws.cell(row=2, column=2))


# ─── Import helpers ────────────────────────────────────────────────────────────

def _cell_str(cell: Any) -> str:
    """Return stripped string value of a cell, or empty string."""
    v = cell.value
    return "" if v is None else str(v).strip()


def _parse_duration(value: Any) -> int | None:
    """Parse duration: '3H30' → 210, '2H' → 120, 90 → 90.

    Returns None if the value cannot be interpreted as a duration.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    m = re.match(r"^(\d+)[Hh](\d+)?$", s)
    if m:
        hours = int(m.group(1))
        mins = int(m.group(2)) if m.group(2) else 0
        return hours * 60 + mins
    try:
        return int(s)
    except ValueError:
        return None


def _parse_unavailable(value: Any) -> list[dict[str, Any]]:
    """Parse unavailability string into slot dicts.

    Formats:
        'mercredi'              → full day off
        'lundi:Matin'           → session off
        'vendredi:08:00-10:00'  → exact time range
    Multiple entries separated by semicolons.
    """
    if not value:
        return []
    slots: list[dict[str, Any]] = []
    for part in str(value).split(";"):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            day, rest = part.split(":", 1)
            day = day.strip()
            rest = rest.strip()
            m = re.match(r"^(\d{1,2}:\d{2})-(\d{1,2}:\d{2})$", rest)
            if m:
                slots.append({"day": day, "start": m.group(1), "end": m.group(2), "session": None})
            else:
                slots.append({"day": day, "start": None, "end": None, "session": rest})
        else:
            slots.append({"day": part, "start": None, "end": None, "session": "all"})
    return slots


# ─── Per-sheet parsers ─────────────────────────────────────────────────────────

def _parse_ecole(wb: Any, errors: list[str]) -> School | None:
    sn = "École"
    if sn not in wb.sheetnames:
        errors.append(f"Onglet « {sn} » introuvable.")
        return None
    ws = wb[sn]
    # Rows 2-4, column B (matching _sheet_ecole layout)
    vals = [_cell_str(ws.cell(row=r, column=2)) for r in range(2, 5)]
    missing = [label for label, v in zip(
        ["Nom de l'école", "Année académique", "Ville"], vals
    ) if not v]
    if missing:
        errors.append(f"École : champs manquants — {', '.join(missing)}.")
        return None
    return School(name=vals[0], academic_year=vals[1], city=vals[2])


def _parse_horaires(wb: Any, errors: list[str]) -> TimeslotConfig | None:
    sn = "Horaires"
    if sn not in wb.sheetnames:
        errors.append(f"Onglet « {sn} » introuvable.")
        return None
    ws = wb[sn]

    # Days: rows 3-9, col A (name), col B (Oui/Non)
    days: list[str] = []
    for r in range(3, 10):
        jour = _cell_str(ws.cell(row=r, column=1))
        actif = _cell_str(ws.cell(row=r, column=2))
        if jour and actif.lower() in ("oui", "yes", "1", "true"):
            days.append(jour.lower())
    if not days:
        errors.append(f"{sn} : aucun jour scolaire actif (lignes 3–9, colonnes A–B).")
        return None

    # Sessions: rows 14-19, cols A-C
    sessions: list[SessionConfig] = []
    for r in range(14, 20):
        nom = _cell_str(ws.cell(row=r, column=1))
        deb = _cell_str(ws.cell(row=r, column=2))
        fin = _cell_str(ws.cell(row=r, column=3))
        if not nom:
            continue
        if not deb or not fin:
            errors.append(f"{sn} ligne {r} : heure manquante pour la session « {nom} ».")
            continue
        sessions.append(SessionConfig(name=nom, start_time=deb, end_time=fin))
    if not sessions:
        errors.append(f"{sn} : aucune session définie (lignes 14–19, colonnes A–C).")
        return None

    # Base unit: row 20, col B
    raw_unit = ws.cell(row=20, column=2).value
    try:
        base_unit = int(raw_unit) if raw_unit is not None else 30
    except (TypeError, ValueError):
        errors.append(f"{sn} ligne 20 : unité de base invalide « {raw_unit} ».")
        base_unit = 30

    return TimeslotConfig.from_simple(day_names=days, sessions=sessions, base_unit_minutes=base_unit)


def _parse_subjects(wb: Any, errors: list[str]) -> list[Subject]:
    sn = "Matières"
    if sn not in wb.sheetnames:
        errors.append(f"Onglet « {sn} » introuvable.")
        return []
    ws = wb[sn]
    subjects: list[Subject] = []
    for r in range(2, ws.max_row + 1):
        nom = _cell_str(ws.cell(row=r, column=1))
        if not nom:
            continue
        abr = _cell_str(ws.cell(row=r, column=2))
        if not abr:
            errors.append(f"Matières ligne {r} : abréviation manquante pour « {nom} ».")
            continue
        color = _cell_str(ws.cell(row=r, column=3)) or "#FFFFFF"
        req_room = _cell_str(ws.cell(row=r, column=4)) or None
        needs_raw = _cell_str(ws.cell(row=r, column=5))
        needs_room = needs_raw.lower() not in ("non", "no", "false", "0")
        subjects.append(Subject(
            name=nom, short_name=abr, color=color,
            required_room_type=req_room, needs_room=needs_room,
        ))
    return subjects


def _parse_teachers(wb: Any, errors: list[str]) -> list[Teacher]:
    sn = "Enseignants"
    if sn not in wb.sheetnames:
        errors.append(f"Onglet « {sn} » introuvable.")
        return []
    ws = wb[sn]
    teachers: list[Teacher] = []
    for r in range(2, ws.max_row + 1):
        nom = _cell_str(ws.cell(row=r, column=1))
        if not nom:
            continue
        matieres_raw = _cell_str(ws.cell(row=r, column=2))
        matieres_seen: set[str] = set()
        matieres: list[str] = []
        for raw_subject in matieres_raw.split(","):
            subject = raw_subject.strip()
            if not subject:
                continue
            key = subject.casefold()
            if key in matieres_seen:
                continue
            matieres_seen.add(key)
            matieres.append(subject)
        if not matieres:
            errors.append(f"Enseignants ligne {r} : matières manquantes pour « {nom} ».")
            continue
        raw_h = ws.cell(row=r, column=3).value
        try:
            max_h = int(raw_h) if raw_h is not None else 20
        except (TypeError, ValueError):
            errors.append(f"Enseignants ligne {r} : heures max invalides « {raw_h} » pour « {nom} ».")
            max_h = 20
        unavail = _parse_unavailable(_cell_str(ws.cell(row=r, column=4)))
        raw_w = ws.cell(row=r, column=5).value
        try:
            weight = float(raw_w) if raw_w is not None else 1.0
        except (TypeError, ValueError):
            weight = 1.0
        teachers.append(Teacher(
            name=nom, subjects=matieres, max_hours_per_week=max_h,
            unavailable_slots=unavail,
        ))
    return teachers


def _parse_classes(wb: Any, errors: list[str]) -> list[SchoolClass]:
    sn = "Classes"
    if sn not in wb.sheetnames:
        errors.append(f"Onglet « {sn} » introuvable.")
        return []
    ws = wb[sn]
    classes: list[SchoolClass] = []
    for r in range(2, ws.max_row + 1):
        nom = _cell_str(ws.cell(row=r, column=1))
        if not nom:
            continue
        niveau = _cell_str(ws.cell(row=r, column=2))
        if not niveau:
            errors.append(f"Classes ligne {r} : niveau manquant pour « {nom} ».")
            continue
        raw_n = ws.cell(row=r, column=3).value
        try:
            count = int(raw_n) if raw_n is not None else 0
        except (TypeError, ValueError):
            errors.append(f"Classes ligne {r} : nombre d'élèves invalide « {raw_n} » pour « {nom} ».")
            count = 0
        classes.append(SchoolClass(name=nom, level=niveau, student_count=count))
    return classes


def _parse_rooms(wb: Any, errors: list[str]) -> list[Room]:
    sn = "Salles"
    if sn not in wb.sheetnames:
        errors.append(f"Onglet « {sn} » introuvable.")
        return []
    ws = wb[sn]
    rooms: list[Room] = []
    for r in range(2, ws.max_row + 1):
        nom = _cell_str(ws.cell(row=r, column=1))
        if not nom:
            continue
        raw_cap = ws.cell(row=r, column=2).value
        try:
            cap = int(raw_cap) if raw_cap is not None else 0
        except (TypeError, ValueError):
            errors.append(f"Salles ligne {r} : capacité invalide « {raw_cap} » pour « {nom} ».")
            cap = 0
        types_raw = _cell_str(ws.cell(row=r, column=3))
        types = [t.strip() for t in types_raw.split(",") if t.strip()]
        rooms.append(Room(name=nom, capacity=cap, types=types))
    return rooms


def _parse_assignments(wb: Any, errors: list[str]) -> list[TeacherAssignment]:
    sn = "Affectations"
    if sn not in wb.sheetnames:
        errors.append(f"Onglet « {sn} » introuvable.")
        return []
    ws = wb[sn]
    assignments: list[TeacherAssignment] = []
    for r in range(2, ws.max_row + 1):
        ens = _cell_str(ws.cell(row=r, column=1))
        mat = _cell_str(ws.cell(row=r, column=2))
        cls = _cell_str(ws.cell(row=r, column=3))
        if not ens and not mat and not cls:
            continue
        missing = [label for label, v in [
            ("Enseignant", ens), ("Matière", mat), ("Classe", cls)
        ] if not v]
        if missing:
            errors.append(
                f"Affectations ligne {r} : colonne(s) manquante(s) — {', '.join(missing)}."
            )
            continue
        assignments.append(TeacherAssignment(teacher=ens, subject=mat, school_class=cls))
    return assignments


def _parse_curriculum(wb: Any, errors: list[str]) -> list[CurriculumEntry]:
    sn = "Programme"
    if sn not in wb.sheetnames:
        errors.append(f"Onglet « {sn} » introuvable.")
        return []
    ws = wb[sn]
    entries: list[CurriculumEntry] = []
    for r in range(2, ws.max_row + 1):
        niveau = _cell_str(ws.cell(row=r, column=1))
        mat = _cell_str(ws.cell(row=r, column=2))
        if not niveau or not mat:
            continue
        total = _parse_duration(ws.cell(row=r, column=3).value)
        if total is None:
            errors.append(
                f"Programme ligne {r} : durée totale invalide « {ws.cell(row=r, column=3).value} »."
            )
            continue
        mode_raw = _cell_str(ws.cell(row=r, column=4)).lower()
        sp_w = _parse_duration(ws.cell(row=r, column=5).value)
        mps = _parse_duration(ws.cell(row=r, column=6).value)
        # Phase 2: min/max session minutes no longer used
        # Derive sessions_per_week and minutes_per_session from total if not provided
        if sp_w is None or mps is None:
            sp_w = sp_w or 1
            mps = mps or total
        entries.append(CurriculumEntry(
            school_class=niveau, subject=mat, total_minutes_per_week=total,
            sessions_per_week=sp_w, minutes_per_session=mps,
        ))
    return entries


def _parse_constraints(wb: Any, errors: list[str]) -> list[Constraint]:
    sn = "Contraintes"
    if sn not in wb.sheetnames:
        errors.append(f"Onglet « {sn} » introuvable.")
        return []
    ws = wb[sn]
    constraints: list[Constraint] = []
    for r in range(2, ws.max_row + 1):
        cid = _cell_str(ws.cell(row=r, column=1))
        if not cid:
            continue
        ctype_raw = _cell_str(ws.cell(row=r, column=2)).lower()
        ctype = "hard" if ctype_raw in ("dur", "hard") else "soft"
        cat = _cell_str(ws.cell(row=r, column=3))
        desc = _cell_str(ws.cell(row=r, column=4))
        raw_prio = ws.cell(row=r, column=5).value
        try:
            prio = int(raw_prio) if raw_prio is not None else 5
        except (TypeError, ValueError):
            prio = 5
        raw_params = _cell_str(ws.cell(row=r, column=6))
        try:
            params: dict = json.loads(raw_params) if raw_params else {}
        except json.JSONDecodeError:
            errors.append(f"Contraintes ligne {r} : JSON invalide pour « {cid} ».")
            params = {}
        constraints.append(Constraint(
            id=cid, type=ctype, category=cat,
            description_fr=desc, priority=prio, parameters=params,
        ))
    return constraints


# ─── Public API ────────────────────────────────────────────────────────────────

def read_template(path: str) -> tuple[SchoolData | None, list[str]]:
    """Parse an Excel template and return a SchoolData object.

    Args:
        path: Path to the .xlsx template file.

    Returns:
        (data, errors) tuple.  data is None when errors are fatal.
        All error strings are in French with sheet name and row number.
    """
    errors: list[str] = []

    try:
        wb = load_workbook(path, data_only=True)
    except FileNotFoundError:
        return None, [f"Fichier introuvable : {path}"]
    except Exception as exc:
        return None, [f"Impossible d'ouvrir le fichier : {exc}"]

    school = _parse_ecole(wb, errors)
    timeslot_config = _parse_horaires(wb, errors)
    subjects = _parse_subjects(wb, errors)
    teachers = _parse_teachers(wb, errors)
    classes = _parse_classes(wb, errors)
    rooms = _parse_rooms(wb, errors)
    assignments = _parse_assignments(wb, errors)
    curriculum = _parse_curriculum(wb, errors)
    constraints = _parse_constraints(wb, errors)

    # Without school and timeslot we cannot build SchoolData at all
    if school is None or timeslot_config is None:
        return None, errors

    # Cross-reference check: Affectations must reference existing entities
    known_teachers = {t.name for t in teachers}
    known_subjects = {s.name for s in subjects}
    known_classes = {c.name for c in classes}
    for i, ta in enumerate(assignments, 2):
        if ta.teacher not in known_teachers:
            errors.append(
                f"Affectations ligne {i} : enseignant « {ta.teacher} » "
                "absent de l'onglet Enseignants."
            )
        if ta.subject not in known_subjects:
            errors.append(
                f"Affectations ligne {i} : matière « {ta.subject} » "
                "absente de l'onglet Matières."
            )
        if ta.school_class not in known_classes:
            errors.append(
                f"Affectations ligne {i} : classe « {ta.school_class} » "
                "absente de l'onglet Classes."
            )

    if errors:
        return None, errors

    data = SchoolData(
        school=school,
        timeslot_config=timeslot_config,
        subjects=subjects,
        teachers=teachers,
        classes=classes,
        rooms=rooms,
        curriculum=curriculum,
        constraints=constraints,
        teacher_assignments=assignments,
    )

    validation_errors = data.validate()
    if validation_errors:
        return None, validation_errors

    return data, []
