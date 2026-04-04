"""Export TimetableResult to a multi-sheet Excel workbook."""
from __future__ import annotations

import logging
from datetime import datetime, timedelta
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from timease.engine.models import Assignment, SchoolData, TimetableResult

logger = logging.getLogger(__name__)

# ─── Styles ────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="CCCCCC")
_FREE_FILL = PatternFill("solid", fgColor="C8E6C9")   # green — teacher free period
_SPACER_FILL = PatternFill("solid", fgColor="F0F0F0")  # between sessions
_FMT = "%H:%M"


def _style_header(cell) -> None:
    cell.fill = _HEADER_FILL
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ─── Time utilities ────────────────────────────────────────────────────────────

def _time_slots(start: str, end: str, base_min: int) -> list[tuple[str, str]]:
    """All (start, end) base-unit intervals between start and end."""
    cur = datetime.strptime(start, _FMT)
    stop = datetime.strptime(end, _FMT)
    delta = timedelta(minutes=base_min)
    slots: list[tuple[str, str]] = []
    while cur < stop:
        nxt = cur + delta
        slots.append((cur.strftime(_FMT), nxt.strftime(_FMT)))
        cur = nxt
    return slots


def _slot_span(start: str, end: str, base_min: int) -> int:
    """Number of base-unit rows an assignment occupies."""
    s = datetime.strptime(start, _FMT)
    e = datetime.strptime(end, _FMT)
    total_min = int((e - s).total_seconds() // 60)
    return max(1, total_min // base_min)


def _duration_min(start: str, end: str) -> int:
    s = datetime.strptime(start, _FMT)
    e = datetime.strptime(end, _FMT)
    return int((e - s).total_seconds() // 60)


# ─── Lookup helpers ────────────────────────────────────────────────────────────

def _build_lookup(
    assignments: list[Assignment],
    perspective: str,
    entity: str,
) -> dict[tuple[str, str], Assignment]:
    """(day, start_time) → Assignment for one entity."""
    filtered = [
        a for a in assignments
        if (perspective == "class" and a.school_class == entity)
        or (perspective == "teacher" and a.teacher == entity)
        or (perspective == "room" and a.room == entity)
    ]
    return {(a.day, a.start_time): a for a in filtered}


def _cell_text(a: Assignment, perspective: str) -> str:
    if perspective == "class":
        parts = [a.subject, a.teacher] + ([a.room] if a.room else [])
    elif perspective == "teacher":
        parts = [a.subject, a.school_class] + ([a.room] if a.room else [])
    else:  # room
        parts = [a.subject, a.school_class, a.teacher]
    return "\n".join(parts)


# ─── Sheet builders ────────────────────────────────────────────────────────────

def _add_summary_sheet(
    wb: Workbook,
    result: TimetableResult,
    data: SchoolData,
) -> None:
    ws = wb.create_sheet("Résumé")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28

    title = ws.cell(row=1, column=1, value=f"Emploi du temps — {data.school.name}")
    title.font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    info_rows = [
        ("Année académique", data.school.academic_year),
        ("Statut", "Résolu ✓" if result.solved else "Non résolu ✗"),
        ("Temps de résolution (s)", f"{result.solve_time_seconds:.2f}"),
        ("Sessions planifiées", len(result.assignments)),
        ("Sessions non planifiées", len(result.unscheduled_sessions)),
        ("Contraintes souples respectées", len(result.soft_constraints_satisfied)),
        ("Contraintes souples violées", len(result.soft_constraints_violated)),
    ]
    for r, (label, val) in enumerate(info_rows, 3):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)

    if result.unscheduled_sessions:
        ws.cell(row=12, column=1, value="Sessions non planifiées :").font = Font(bold=True)
        for i, sess in enumerate(result.unscheduled_sessions, 13):
            ws.cell(row=i, column=1, value=str(sess))

    if result.soft_constraints_violated:
        start_row = 13 + len(result.unscheduled_sessions) + 1
        ws.cell(row=start_row, column=1,
                value="Contraintes souples violées :").font = Font(bold=True)
        for i, cid in enumerate(result.soft_constraints_violated, start_row + 1):
            ws.cell(row=i, column=1, value=cid)


def _add_entity_sheet(
    wb: Workbook,
    result: TimetableResult,
    data: SchoolData,
    perspective: str,
    entity: str,
) -> None:
    prefix = {"class": "Classe", "teacher": "Prof", "room": "Salle"}[perspective]
    sheet_name = f"{prefix} — {entity}"[:31]  # Excel enforces max 31 chars
    ws = wb.create_sheet(sheet_name)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    days = data.timeslot_config.days
    base_min = data.timeslot_config.base_unit_minutes
    subject_colors = {s.name: s.color.lstrip("#") for s in data.subjects}
    lookup = _build_lookup(result.assignments, perspective, entity)
    is_teacher = perspective == "teacher"

    # Column widths
    ws.column_dimensions["A"].width = 13
    for col in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 19

    # Row 1 — day headers
    _style_header(ws.cell(row=1, column=1, value="Heure"))
    for col, day in enumerate(days, 2):
        _style_header(ws.cell(row=1, column=col, value=day.capitalize()))
    ws.freeze_panes = "B2"

    # Build ordered list of grid rows: (session_name, start, end) or None (spacer)
    grid_rows: list[tuple[str, str, str] | None] = []
    for session in data.timeslot_config.sessions:
        for s, e in _time_slots(session.start_time, session.end_time, base_min):
            grid_rows.append((session.name, s, e))
        grid_rows.append(None)
    # Remove trailing spacer
    while grid_rows and grid_rows[-1] is None:
        grid_rows.pop()

    # Fill grid
    current_row = 2
    # col → last row already covered by a vertical merge
    merged_until: dict[int, int] = {}

    for slot in grid_rows:
        if slot is None:
            # Inter-session spacer
            ws.row_dimensions[current_row].height = 5
            for col in range(1, len(days) + 2):
                ws.cell(row=current_row, column=col).fill = _SPACER_FILL
            current_row += 1
            continue

        _, start, end = slot
        t_cell = ws.cell(row=current_row, column=1, value=f"{start}–{end}")
        t_cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.row_dimensions[current_row].height = 32

        for col, day in enumerate(days, 2):
            if merged_until.get(col, 0) >= current_row:
                continue  # already consumed by a previous vertical merge

            key = (day, start)
            if key in lookup:
                a = lookup[key]
                span = _slot_span(a.start_time, a.end_time, base_min)
                color = subject_colors.get(a.subject, "FFFFFF")

                cell = ws.cell(row=current_row, column=col, value=_cell_text(a, perspective))
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(
                    wrap_text=True, vertical="center", horizontal="center"
                )
                cell.font = Font(size=9)

                if span > 1:
                    ws.merge_cells(
                        start_row=current_row, start_column=col,
                        end_row=current_row + span - 1, end_column=col,
                    )
                    merged_until[col] = current_row + span - 1
            elif is_teacher:
                ws.cell(row=current_row, column=col).fill = _FREE_FILL

        current_row += 1


def _add_subject_summary(
    wb: Workbook,
    result: TimetableResult,
    data: SchoolData,
) -> None:
    ws = wb.create_sheet("Par matière")
    ws.column_dimensions["A"].width = 26

    classes = [c.name for c in data.classes]
    subjects = [s.name for s in data.subjects]
    subject_colors = {s.name: s.color.lstrip("#") for s in data.subjects}

    # Header row
    _style_header(ws.cell(row=1, column=1, value="Matière"))
    for col, cls in enumerate(classes, 2):
        _style_header(ws.cell(row=1, column=col, value=cls))
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = "B2"

    # Accumulate scheduled minutes per (subject, class)
    minutes: dict[tuple[str, str], int] = {}
    for a in result.assignments:
        key = (a.subject, a.school_class)
        minutes[key] = minutes.get(key, 0) + _duration_min(a.start_time, a.end_time)

    for row, subj in enumerate(subjects, 2):
        color = subject_colors.get(subj, "FFFFFF")
        subj_cell = ws.cell(row=row, column=1, value=subj)
        subj_cell.fill = PatternFill("solid", fgColor=color)
        subj_cell.font = Font(bold=True)
        for col, cls in enumerate(classes, 2):
            mins = minutes.get((subj, cls), 0)
            if mins:
                cell = ws.cell(row=row, column=col, value=mins)
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal="center")


# ─── Public API ────────────────────────────────────────────────────────────────

def export_timetable(
    result: TimetableResult,
    data: SchoolData,
    output_path: str,
    perspectives: list[str] | None = None,
) -> None:
    """Export a solved timetable to a multi-sheet Excel workbook.

    Args:
        result:       Solved TimetableResult from the engine.
        data:         Original SchoolData used for the solve.
        output_path:  Destination .xlsx path.
        perspectives: Which views to include.  Defaults to all four:
                      ['class', 'teacher', 'room', 'subject'].
    """
    if perspectives is None:
        perspectives = ["class", "teacher", "room", "subject"]

    wb = Workbook()
    wb.remove(wb.active)

    _add_summary_sheet(wb, result, data)

    if "class" in perspectives:
        for cls in data.classes:
            _add_entity_sheet(wb, result, data, "class", cls.name)

    if "teacher" in perspectives:
        for teacher in data.teachers:
            _add_entity_sheet(wb, result, data, "teacher", teacher.name)

    if "room" in perspectives:
        for room in data.rooms:
            _add_entity_sheet(wb, result, data, "room", room.name)

    if "subject" in perspectives:
        _add_subject_summary(wb, result, data)

    wb.save(output_path)
    logger.info("Export Excel : %s (%d onglets)", output_path, len(wb.sheetnames))
