"""
Tests for timease/io modules: import, export (Excel/PDF/Word/Markdown),
and file_parser.

All export tests share one module-scoped solver run to keep the suite fast.
"""

from __future__ import annotations

import dataclasses
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from timease.engine.models import SchoolData

DAKAR_JSON = (
    Path(__file__).parent.parent / "timease" / "data" / "real_school_dakar_LOCKED.json"
)
TEMPLATE_XLSX = (
    Path(__file__).parent.parent / "timease" / "data" / "template.xlsx"
)


# ---------------------------------------------------------------------------
# Module-scoped fixtures  (solver runs once for the whole file)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def dakar_data() -> SchoolData:
    return SchoolData.from_json(DAKAR_JSON)


@pytest.fixture(scope="module")
def dakar_result(dakar_data: SchoolData):
    from timease.engine.solver import TimetableSolver
    return TimetableSolver().solve(dakar_data, timeout_seconds=120)


# ---------------------------------------------------------------------------
# Template tests
# ---------------------------------------------------------------------------

def test_template_has_10_sheets(tmp_path: Path) -> None:
    """create_template() produces exactly 10 sheets."""
    from timease.io.excel_import import create_template

    xlsx = tmp_path / "template.xlsx"
    create_template(str(xlsx))
    wb = load_workbook(str(xlsx))
    assert len(wb.sheetnames) == 10, f"Got sheets: {wb.sheetnames}"


def test_template_affectations_headers(tmp_path: Path) -> None:
    """Affectations sheet has the three expected column headers."""
    from timease.io.excel_import import create_template

    xlsx = tmp_path / "template.xlsx"
    create_template(str(xlsx))
    wb = load_workbook(str(xlsx))
    ws = wb["Affectations"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
    assert headers == ["Enseignant *", "Matière *", "Classe *"]


# ---------------------------------------------------------------------------
# Import tests
# ---------------------------------------------------------------------------

def test_import_strips_whitespace() -> None:
    """_cell_str strips leading/trailing whitespace from cell values."""
    from timease.io.excel_import import _cell_str

    class _FakeCell:
        def __init__(self, value: object) -> None:
            self.value = value

    assert _cell_str(_FakeCell("  Mme Diallo  ")) == "Mme Diallo"
    assert _cell_str(_FakeCell("\t6ème A\n"))      == "6ème A"
    assert _cell_str(_FakeCell("  "))              == ""
    assert _cell_str(_FakeCell(None))              == ""


def test_import_parses_3h30_as_210min() -> None:
    """_parse_duration correctly converts duration strings and numbers."""
    from timease.io.excel_import import _parse_duration

    assert _parse_duration("3H30") == 210
    assert _parse_duration("3h30") == 210
    assert _parse_duration("2H")   == 120
    assert _parse_duration("2h")   == 120
    assert _parse_duration(210)    == 210
    assert _parse_duration(90)     == 90
    assert _parse_duration("90")   == 90
    assert _parse_duration(None)   is None
    assert _parse_duration("")     is None


def test_import_rejects_unknown_teacher(tmp_path: Path) -> None:
    """read_template returns errors when Affectations references an unknown teacher."""
    from timease.io.excel_import import create_template, read_template

    xlsx = tmp_path / "bad.xlsx"
    create_template(str(xlsx))

    # Overwrite the example Affectations row with an unknown teacher
    wb = load_workbook(str(xlsx))
    ws = wb["Affectations"]
    ws.cell(row=2, column=1, value="Enseignant Fantôme")
    ws.cell(row=2, column=2, value="Mathématiques")
    ws.cell(row=2, column=3, value="6ème A")
    wb.save(str(xlsx))

    data, errors = read_template(str(xlsx))

    assert data is None, "read_template should return None on unknown teacher"
    assert errors, "errors list should be non-empty"
    combined = " ".join(errors)
    assert "Fantôme" in combined or "Enseignant Fantôme" in combined, (
        f"Expected unknown-teacher error, got: {errors}"
    )


def test_import_teacher_subjects_are_deduplicated_case_insensitive(tmp_path: Path) -> None:
    """Teacher subjects from Excel should keep one normalized entry per label."""
    from timease.io.excel_import import create_template, read_template

    xlsx = tmp_path / "dedup-subjects.xlsx"
    create_template(str(xlsx))

    wb = load_workbook(str(xlsx))
    ws_teachers = wb["Enseignants"]
    ws_teachers.cell(row=2, column=1, value="Mme Test")
    ws_teachers.cell(row=2, column=2, value="Maths, maths,  MATHS  , Physique")
    ws_teachers.cell(row=2, column=3, value=18)

    ws_subjects = wb["Matières"]
    ws_subjects.cell(row=2, column=1, value="Maths")
    ws_subjects.cell(row=2, column=2, value="M")
    ws_subjects.cell(row=3, column=1, value="Physique")
    ws_subjects.cell(row=3, column=2, value="P")

    ws_classes = wb["Classes"]
    ws_classes.cell(row=2, column=1, value="6ème A")
    ws_classes.cell(row=2, column=2, value="6ème")
    ws_classes.cell(row=2, column=3, value=30)

    ws_assign = wb["Affectations"]
    ws_assign.cell(row=2, column=1, value="Mme Test")
    ws_assign.cell(row=2, column=2, value="Maths")
    ws_assign.cell(row=2, column=3, value="6ème A")

    ws_curr = wb["Programme"]
    ws_curr.cell(row=2, column=1, value="6ème A")
    ws_curr.cell(row=2, column=2, value="Maths")
    ws_curr.cell(row=2, column=3, value=60)
    ws_curr.cell(row=2, column=4, value=1)
    ws_curr.cell(row=2, column=5, value=60)

    wb.save(str(xlsx))

    data, errors = read_template(str(xlsx))
    assert not errors, errors
    assert data is not None
    assert data.teachers[0].subjects == ["Maths", "Physique"]


# ---------------------------------------------------------------------------
# Export tests  (reuse module-scoped dakar_result)
# ---------------------------------------------------------------------------

def test_export_excel_has_all_perspectives(dakar_result, dakar_data, tmp_path: Path) -> None:
    """Default Excel export includes class, teacher, room, and subject sheets."""
    from timease.io.excel_export import export_timetable

    xlsx = tmp_path / "out.xlsx"
    export_timetable(dakar_result, dakar_data, str(xlsx))

    wb = load_workbook(str(xlsx))
    names = wb.sheetnames

    assert "Résumé" in names
    assert any("Classe" in n for n in names),    f"No class sheet in {names}"
    assert any("Prof"   in n for n in names),    f"No teacher sheet in {names}"
    assert any("Salle"  in n for n in names),    f"No room sheet in {names}"
    assert "Par matière" in names,               f"No subject summary in {names}"


def test_export_excel_teacher_colors_not_subject_colors(dakar_result, dakar_data, tmp_path: Path) -> None:
    """Class sheet cells should use teacher color coding."""
    from timease.io.excel_export import export_timetable
    from timease.utils.teacher_colors import teacher_color

    xlsx = tmp_path / "teacher-colors.xlsx"
    export_timetable(dakar_result, dakar_data, str(xlsx), perspectives=["class"])
    wb = load_workbook(str(xlsx))
    class_sheet_name = next(n for n in wb.sheetnames if n.startswith("Classe"))
    ws = wb[class_sheet_name]

    # Find first scheduled cell on class sheet
    target = None
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and "\n" in val:
                target = ws.cell(row=row, column=col)
                break
        if target is not None:
            break

    assert target is not None, "No scheduled cell found on class sheet"
    lines = str(target.value).split("\n")
    assert len(lines) >= 2
    teacher = lines[1]
    expected = teacher_color(teacher).lstrip("#").upper()
    actual = target.fill.fgColor.rgb[-6:].upper() if target.fill.fgColor and target.fill.fgColor.rgb else ""
    assert actual == expected


def test_export_maguette_both_subjects(dakar_result, dakar_data, tmp_path: Path) -> None:
    """Markdown export contains both subjects that Maguette teaches."""
    from timease.io.md_export import export_markdown

    md_path = tmp_path / "out.md"
    export_markdown(dakar_result, dakar_data, str(md_path))
    content = md_path.read_text(encoding="utf-8")

    maguette_subjects = {
        a.subject for a in dakar_result.assignments if a.teacher == "Maguette"
    }
    assert len(maguette_subjects) >= 2, (
        f"Expected Maguette to teach ≥2 subjects in the result, "
        f"got: {maguette_subjects}"
    )
    assert "Maguette" in content, "Teacher name 'Maguette' missing from markdown"
    for subject in maguette_subjects:
        assert subject in content, f"Subject '{subject}' missing from markdown"


def test_export_pdf_creates_file(dakar_result, dakar_data, tmp_path: Path) -> None:
    """export_pdf writes a non-empty .pdf file."""
    from timease.io.pdf_export import export_pdf

    pdf = tmp_path / "out.pdf"
    export_pdf(dakar_result, dakar_data, str(pdf))

    assert pdf.exists(), "PDF file was not created"
    assert pdf.stat().st_size > 0, "PDF file is empty"


def test_export_word_creates_file(dakar_result, dakar_data, tmp_path: Path) -> None:
    """export_word writes a non-empty .docx file."""
    from timease.io.word_export import export_word

    docx = tmp_path / "out.docx"
    export_word(dakar_result, dakar_data, str(docx))

    assert docx.exists(), "DOCX file was not created"
    assert docx.stat().st_size > 0, "DOCX file is empty"


def test_export_markdown_has_all_classes(dakar_result, dakar_data, tmp_path: Path) -> None:
    """Markdown class-perspective export contains a section for every class."""
    from timease.io.md_export import export_markdown

    md = tmp_path / "out.md"
    export_markdown(dakar_result, dakar_data, str(md), perspectives=["class"])
    content = md.read_text(encoding="utf-8")

    for cls in dakar_data.classes:
        assert cls.name in content, f"Class '{cls.name}' missing from markdown"


def test_api_estimation_endpoint_exists() -> None:
    from fastapi.testclient import TestClient
    from timease.api.main import app

    client = TestClient(app)
    sid = client.post("/api/session").json()["session_id"]
    est = client.get(f"/api/session/{sid}/solve-estimate")
    assert est.status_code == 200
    data = est.json()
    assert "suggested_timeout_seconds" in data


def test_api_estimation_no_room_pressure_when_no_rooms() -> None:
    from fastapi.testclient import TestClient
    from timease.api.main import app

    client = TestClient(app)
    sid = client.post("/api/session").json()["session_id"]
    client.put(
        f"/api/session/{sid}/school_data",
        json={
            "classes": [{"name": "6e A", "level": "Collège", "student_count": 30}],
            "rooms": [],
            "curriculum": [],
            "constraints": [],
            "days": [],
            "teachers": [],
        },
    )
    est = client.get(f"/api/session/{sid}/solve-estimate")
    assert est.status_code == 200
    factors = est.json().get("factors", [])
    assert any("aucune salle définie" in f for f in factors)
    assert not any("pression sur les salles" in f for f in factors)


def test_api_solve_mode_applies_timeout_cap_and_fast_flags(dakar_data: SchoolData, monkeypatch: pytest.MonkeyPatch) -> None:
    from fastapi.testclient import TestClient
    from timease.api.main import app
    from timease.engine.models import SchoolData as EngineSchoolData, TimetableResult
    from timease.engine.solver import TimetableSolver

    seen: dict[str, int] = {}

    seen_flags: dict[str, bool] = {}

    def fake_solve(
        self,
        data,  # noqa: ANN001
        timeout_seconds=30,
        optimize_soft_constraints=True,
        stop_at_first_solution=False,
        enforce_room_conflicts=True,
    ):
        seen["timeout"] = int(timeout_seconds)
        seen_flags["optimize_soft_constraints"] = bool(optimize_soft_constraints)
        seen_flags["stop_at_first_solution"] = bool(stop_at_first_solution)
        seen_flags["enforce_room_conflicts"] = bool(enforce_room_conflicts)
        return TimetableResult(
            assignments=[],
            solved=False,
            solve_time_seconds=float(max(1, timeout_seconds - 1)),
            conflicts=[{"reason": "UNKNOWN"}],
        )

    monkeypatch.setattr(TimetableSolver, "solve", fake_solve)
    monkeypatch.setattr(EngineSchoolData, "validate", lambda self: [])

    client = TestClient(app)
    sid = client.post("/api/session").json()["session_id"]
    client.post(
        f"/api/session/{sid}/restore",
        json={
            "school_data": dataclasses.asdict(dakar_data),
            "teacher_assignments": [
                {
                    "teacher": ta.teacher,
                    "subject": ta.subject,
                    "school_class": ta.school_class,
                }
                for ta in dakar_data.teacher_assignments
            ],
        },
    )

    res = client.post(
        f"/api/session/{sid}/solve",
        json={"timeout": 10, "mode": "complete", "request_id": "test-mode-floor"},
    )
    assert res.status_code == 200
    body = res.json()
    assert body["status"] == "TIMEOUT"
    assert seen["timeout"] == 30
    assert body["effective_timeout_seconds"] == 30
    assert seen_flags["optimize_soft_constraints"] is True
    assert seen_flags["stop_at_first_solution"] is False

    res_fast = client.post(
        f"/api/session/{sid}/solve",
        json={"mode": "fast", "request_id": "test-fast-mode"},
    )
    assert res_fast.status_code == 200
    body_fast = res_fast.json()
    assert body_fast["status"] == "TIMEOUT"
    assert seen["timeout"] <= 60
    assert body_fast["effective_timeout_seconds"] <= 60
    assert seen_flags["optimize_soft_constraints"] is False
    assert seen_flags["stop_at_first_solution"] is True
    assert seen_flags["enforce_room_conflicts"] is False


def test_api_job_lifecycle_create_cancel_delete(dakar_data: SchoolData) -> None:
    from fastapi.testclient import TestClient
    from timease.api.main import app

    client = TestClient(app)
    sid = client.post("/api/session").json()["session_id"]

    snapshot = client.post(
        f"/api/session/{sid}/snapshots",
        json={
            "name": "Version test",
            "school_data": dataclasses.asdict(dakar_data),
            "teacher_assignments": [
                {"teacher": ta.teacher, "subject": ta.subject, "school_class": ta.school_class}
                for ta in dakar_data.teacher_assignments
            ],
        },
    )
    assert snapshot.status_code == 200
    snap_id = snapshot.json()["snapshot"]["id"]

    created = client.post(
        f"/api/session/{sid}/jobs",
        json={"snapshot_id": snap_id, "mode": "complete", "request_id": "job-lifecycle"},
    )
    assert created.status_code == 200
    job_id = created.json()["job"]["id"]

    cancelled = client.post(f"/api/session/{sid}/jobs/{job_id}/cancel")
    assert cancelled.status_code == 200
    assert cancelled.json()["job"]["status"] == "cancelled"
    report = cancelled.json()["job"].get("report")
    assert isinstance(report, dict)
    assert report["reason_code"] == "CANCELLED_BY_USER"
    assert "arrêtée" in report["reason_message"]

    deleted = client.delete(f"/api/session/{sid}/jobs/{job_id}")
    assert deleted.status_code == 200
    assert deleted.json()["ok"] is True

    listed = client.get(f"/api/session/{sid}/jobs")
    assert listed.status_code == 200
    assert all(j["id"] != job_id for j in listed.json().get("jobs", []))


def test_api_snapshot_delete_removes_snapshot_and_related_finished_jobs(dakar_data: SchoolData) -> None:
    from fastapi.testclient import TestClient
    from timease.api.main import app

    client = TestClient(app)
    sid = client.post("/api/session").json()["session_id"]

    snapshot = client.post(
        f"/api/session/{sid}/snapshots",
        json={
            "name": "Version à supprimer",
            "school_data": dataclasses.asdict(dakar_data),
            "teacher_assignments": [
                {"teacher": ta.teacher, "subject": ta.subject, "school_class": ta.school_class}
                for ta in dakar_data.teacher_assignments
            ],
        },
    )
    assert snapshot.status_code == 200
    snap_id = snapshot.json()["snapshot"]["id"]

    created = client.post(
        f"/api/session/{sid}/jobs",
        json={"snapshot_id": snap_id, "mode": "fast", "request_id": "snapshot-delete"},
    )
    assert created.status_code == 200
    job_id = created.json()["job"]["id"]

    cancelled = client.post(f"/api/session/{sid}/jobs/{job_id}/cancel")
    assert cancelled.status_code == 200

    deleted = client.delete(f"/api/session/{sid}/snapshots/{snap_id}")
    assert deleted.status_code == 200
    body = deleted.json()
    assert body["ok"] is True
    assert body["deleted_jobs"] >= 1

    snaps = client.get(f"/api/session/{sid}/snapshots").json().get("snapshots", [])
    assert all(s["id"] != snap_id for s in snaps)


def test_api_snapshot_delete_blocks_when_job_running(dakar_data: SchoolData) -> None:
    from fastapi.testclient import TestClient
    from timease.api.main import app, sessions

    client = TestClient(app)
    sid = client.post("/api/session").json()["session_id"]

    snapshot = client.post(
        f"/api/session/{sid}/snapshots",
        json={
            "name": "Version occupée",
            "school_data": dataclasses.asdict(dakar_data),
            "teacher_assignments": [
                {"teacher": ta.teacher, "subject": ta.subject, "school_class": ta.school_class}
                for ta in dakar_data.teacher_assignments
            ],
        },
    )
    assert snapshot.status_code == 200
    snap_id = snapshot.json()["snapshot"]["id"]

    # Inject a deterministic running job linked to this snapshot.
    sessions[sid].setdefault("jobs", []).append(
        {
            "id": "running-delete-block",
            "snapshot_id": snap_id,
            "status": "running",
            "mode": "complete",
            "request_id": "snapshot-delete-block",
            "created_at": 0.0,
            "started_at": 0.0,
            "finished_at": None,
            "estimate": {},
            "result": None,
            "report": None,
        }
    )

    blocked = client.delete(f"/api/session/{sid}/snapshots/{snap_id}")
    assert blocked.status_code == 409
    assert "job en cours" in blocked.json()["detail"]


def test_api_poll_jobs_sets_report_when_worker_returns_no_payload(dakar_data: SchoolData) -> None:
    from fastapi.testclient import TestClient
    from timease.api.main import app, sessions, _poll_jobs

    class DeadProcess:
        def is_alive(self) -> bool:
            return False

    class EmptyQueue:
        def empty(self) -> bool:
            return True

    client = TestClient(app)
    sid = client.post("/api/session").json()["session_id"]

    sessions[sid]["jobs"] = [
        {
            "id": "job-no-payload",
            "snapshot_id": "snap-x",
            "status": "running",
            "mode": "balanced",
            "request_id": "req-no-payload",
            "effective_timeout_seconds": 180,
            "created_at": 0.0,
            "started_at": 0.0,
            "finished_at": None,
            "estimate": {},
            "result": None,
            "report": None,
        }
    ]

    from timease.api import main as api_main
    api_main.job_runtime_handles["job-no-payload"] = {
        "process": DeadProcess(),
        "queue": EmptyQueue(),
        "sid": sid,
    }

    _poll_jobs(sid)
    job = sessions[sid]["jobs"][0]
    assert job["status"] == "failed"
    assert job.get("report", {}).get("reason_code") == "WORKER_NO_RESULT"
    assert "sans résultat" in job.get("report", {}).get("reason_message", "")


def test_api_snapshot_rename_happy_path_and_validation(dakar_data: SchoolData) -> None:
    from fastapi.testclient import TestClient
    from timease.api.main import app

    client = TestClient(app)
    sid = client.post("/api/session").json()["session_id"]

    snapshot = client.post(
        f"/api/session/{sid}/snapshots",
        json={
            "name": "Version initiale",
            "school_data": dataclasses.asdict(dakar_data),
            "teacher_assignments": [
                {"teacher": ta.teacher, "subject": ta.subject, "school_class": ta.school_class}
                for ta in dakar_data.teacher_assignments
            ],
        },
    )
    assert snapshot.status_code == 200
    snap_id = snapshot.json()["snapshot"]["id"]

    renamed = client.patch(
        f"/api/session/{sid}/snapshots/{snap_id}",
        json={"name": "Version renommée"},
    )
    assert renamed.status_code == 200
    assert renamed.json()["snapshot"]["name"] == "Version renommée"

    invalid = client.patch(
        f"/api/session/{sid}/snapshots/{snap_id}",
        json={"name": "   "},
    )
    assert invalid.status_code == 400
    assert "nom de la version" in invalid.json()["detail"].lower()


# ---------------------------------------------------------------------------
# File parser tests
# ---------------------------------------------------------------------------

def test_file_parser_excel() -> None:
    """extract_content on the bundled template returns type='excel' and sheet names."""
    from timease.io.file_parser import extract_content

    content, ftype = extract_content(str(TEMPLATE_XLSX))

    assert ftype == "excel"
    assert "Affectations" in content, (
        "'Affectations' sheet name missing from parsed Excel content"
    )


def test_file_parser_unknown_format(tmp_path: Path) -> None:
    """extract_content returns ('Format non supporté', 'unknown') for unsupported extensions."""
    from timease.io.file_parser import extract_content

    fake = tmp_path / "data.xyz"
    fake.write_text("dummy content")

    content, ftype = extract_content(str(fake))

    assert ftype == "unknown"
    assert "non supporté" in content.lower(), (
        f"Expected 'non supporté' in content, got: {content!r}"
    )
